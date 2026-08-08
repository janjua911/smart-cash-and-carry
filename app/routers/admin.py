from decimal import Decimal, InvalidOperation
from datetime import datetime  # ✅ Added for bulk import
import io  # ✅ Added for bulk import

from fastapi import APIRouter, Depends, HTTPException, Request, status, UploadFile, File  # ✅ Added UploadFile, File
from fastapi.responses import RedirectResponse
from sqlalchemy import func, select
from sqlalchemy.orm import Session, joinedload

from app.core.security import csrf_is_valid, get_csrf_token, set_flash, verify_password
from app.database import get_db
from app.dependencies import require_admin
from app.models.admin import Admin
from app.models.category import Category
from app.models.order import Order, OrderStatus
from app.models.product import Product, StockStatus
from app.models.variant import ProductVariant
from app.services.slug import unique_category_slug, unique_product_slug
from app.services.storage import delete_local_image, save_product_image
from app.web import template_context, templates


router = APIRouter(prefix="/admin", tags=["admin"])


def verify_csrf(request: Request, form) -> None:
    if not csrf_is_valid(request, form.get("csrf_token")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid form token")


def product_form_values(form) -> dict:
    return {
        "name": str(form.get("name", "")).strip(),
        "slug": str(form.get("slug", "")).strip(),
        "description": str(form.get("description", "")).strip(),
        "unit": str(form.get("unit", "1 item")).strip(),
        "price": str(form.get("price", "")).strip(),
        "compare_at_price": str(form.get("compare_at_price", "")).strip(),
        "stock_quantity": str(form.get("stock_quantity", "0")).strip(),
        "stock_status": str(form.get("stock_status", StockStatus.IN_STOCK.value)),
        "category_id": str(form.get("category_id", "")),
        "is_active": form.get("is_active") == "on",
        "is_featured": form.get("is_featured") == "on",
    }


async def apply_product_form(db: Session, product: Product, form) -> None:
    values = product_form_values(form)
    if len(values["name"]) < 2:
        raise ValueError("Product name must contain at least 2 characters.")
    try:
        category_id = int(values["category_id"])
        stock_quantity = max(0, int(values["stock_quantity"]))
        price = Decimal(values["price"])
        compare_at_price = (
            Decimal(values["compare_at_price"]) if values["compare_at_price"] else None
        )
        stock_status = values["stock_status"].strip().lower()
    except (ValueError, InvalidOperation) as exc:
        raise ValueError("Please provide valid price, stock and category values.") from exc

    if price <= 0:
        raise ValueError("Product price must be greater than zero.")
    if compare_at_price is not None and compare_at_price <= price:
        raise ValueError("Compare-at price must be higher than the selling price.")
    if db.get(Category, category_id) is None:
        raise ValueError("Selected category does not exist.")
    
    # Stock status handling - explicit fix
    if stock_status not in {"in_stock", "out_of_stock", "sold_out"}:
        raise ValueError("Invalid stock status.")
    product.stock_status = StockStatus(stock_status)

    # Is_active should be independent of stock_status
    is_active_val = values["is_active"]
    product.is_active = is_active_val in {"on", "true", "1", "yes", True}

    # Auto-update stock_status based on quantity if explicitly set
    stock_qty = values["stock_quantity"]
    if stock_qty:
        try:
            qty = int(stock_qty)
            product.stock_quantity = max(0, qty)
            # If stock is 0 and status is IN_STOCK, auto-set to OUT_OF_STOCK
            if qty == 0 and product.stock_status == StockStatus.IN_STOCK:
                product.stock_status = StockStatus.OUT_OF_STOCK
            elif qty > 0 and product.stock_status in (StockStatus.OUT_OF_STOCK, StockStatus.SOLD_OUT):
                # If stock > 0 and was out_of_stock, make it in_stock
                product.stock_status = StockStatus.IN_STOCK
        except ValueError:
            raise ValueError("Stock quantity must be a number.")

    image_upload = form.get("image")
    new_image_path = None
    if getattr(image_upload, "filename", ""):
        new_image_path = await save_product_image(image_upload)

    old_image_path = product.image_path
    product.category_id = category_id
    product.name = values["name"]
    product.slug = unique_product_slug(
        db,
        values["slug"] or values["name"],
        product.id,
    )
    product.description = values["description"]
    product.unit = values["unit"] or "1 item"
    product.price = price
    product.compare_at_price = compare_at_price
    product.is_featured = values["is_featured"]
    if new_image_path:
        product.image_path = new_image_path
        delete_local_image(old_image_path)


# Helper function to handle variants
def handle_product_variants(db: Session, product_id: int, form) -> None:
    """Create/update/delete variants for a product."""
    variant_names = form.getlist("variant_name[]")
    variant_prices = form.getlist("variant_price[]")
    variant_stocks = form.getlist("variant_stock[]")
    variant_ids = form.getlist("variant_id[]")  # For editing existing variants
    
    # If no variant data, skip
    if not variant_names or not any(name.strip() for name in variant_names):
        return
    
    # Delete existing variants if editing
    if product_id:
        db.query(ProductVariant).filter(ProductVariant.product_id == product_id).delete()
    
    # Create new variants
    for i, name in enumerate(variant_names):
        if name.strip():
            # Get price adjustment (handle empty or invalid values)
            price_adj_str = variant_prices[i] if i < len(variant_prices) else "0"
            stock_str = variant_stocks[i] if i < len(variant_stocks) else "0"
            
            try:
                price_adjustment = float(price_adj_str.strip()) if price_adj_str.strip() else 0
            except ValueError:
                price_adjustment = 0
            
            try:
                stock_quantity = int(stock_str.strip()) if stock_str.strip() else 0
            except ValueError:
                stock_quantity = 0
            
            variant = ProductVariant(
                product_id=product_id,
                name=name.strip(),
                price_adjustment=price_adjustment,
                stock_quantity=max(0, stock_quantity),
                is_active=True,  # New variants are active by default
            )
            db.add(variant)


@router.get("/login")
def login_page(request: Request):
    if request.session.get("admin_id"):
        return RedirectResponse("/admin", status_code=status.HTTP_303_SEE_OTHER)
    return templates.TemplateResponse(
        request=request,
        name="admin/login.html",
        context={"request": request, "csrf_token": get_csrf_token(request), "error": None},
    )


@router.post("/login")
async def login(request: Request, db: Session = Depends(get_db)):
    form = await request.form()
    verify_csrf(request, form)
    email = str(form.get("email", "")).strip().lower()
    password = str(form.get("password", ""))
    admin = db.scalar(select(Admin).where(Admin.email == email))
    if not admin or not admin.is_active or not verify_password(password, admin.password_hash):
        return templates.TemplateResponse(
            request=request,
            name="admin/login.html",
            context={
                "request": request,
                "csrf_token": get_csrf_token(request),
                "error": "Invalid email or password.",
                "email": email,
            },
            status_code=status.HTTP_401_UNAUTHORIZED,
        )
    request.session.clear()
    request.session["admin_id"] = admin.id
    get_csrf_token(request)
    return RedirectResponse("/admin", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/logout")
async def logout(request: Request):
    form = await request.form()
    verify_csrf(request, form)
    request.session.clear()
    return RedirectResponse("/admin/login", status_code=status.HTTP_303_SEE_OTHER)


@router.get("")
def dashboard(
    request: Request,
    db: Session = Depends(get_db),
    admin: Admin = Depends(require_admin),
):
    stats = {
        "products": db.scalar(select(func.count()).select_from(Product)) or 0,
        "active_products": db.scalar(
            select(func.count()).select_from(Product).where(Product.is_active.is_(True))
        )
        or 0,
        "low_stock": db.scalar(
            select(func.count())
            .select_from(Product)
            .where(Product.is_active.is_(True), Product.stock_quantity <= 5)
        )
        or 0,
        "orders": db.scalar(select(func.count()).select_from(Order)) or 0,
        "pending_orders": db.scalar(
            select(func.count()).select_from(Order).where(Order.status == OrderStatus.PENDING)
        )
        or 0,
        "revenue": db.scalar(
            select(func.coalesce(func.sum(Order.total_amount), 0)).where(
                Order.status != OrderStatus.CANCELLED
            )
        )
        or 0,
    }
    recent_orders = db.scalars(
        select(Order).order_by(Order.created_at.desc()).limit(8)
    ).all()
    low_stock_products = db.scalars(
        select(Product)
        .where(Product.is_active.is_(True), Product.stock_quantity <= 5)
        .order_by(Product.stock_quantity)
        .limit(8)
    ).all()
    return templates.TemplateResponse(
        request=request,
        name="admin/dashboard.html",
        context=template_context(
            request,
            admin=admin,
            stats=stats,
            recent_orders=recent_orders,
            low_stock_products=low_stock_products,
        ),
    )


@router.get("/products")
def products_list(
    request: Request,
    db: Session = Depends(get_db),
    admin: Admin = Depends(require_admin),
):
    products = db.scalars(
        select(Product)
        .options(joinedload(Product.category))
        .order_by(Product.created_at.desc())
    ).unique().all()
    return templates.TemplateResponse(
        request=request,
        name="admin/products/list.html",
        context=template_context(request, admin=admin, products=products),
    )


@router.get("/products/new")
def product_new_page(
    request: Request,
    db: Session = Depends(get_db),
    admin: Admin = Depends(require_admin),
):
    categories = db.scalars(select(Category).where(Category.is_active.is_(True)).order_by(Category.name)).all()
    return templates.TemplateResponse(
        request=request,
        name="admin/products/form.html",
        context=template_context(
            request, admin=admin, product=None, categories=categories, form_values={}, error=None
        ),
    )


@router.post("/products/new")
async def product_create(
    request: Request,
    db: Session = Depends(get_db),
    admin: Admin = Depends(require_admin),
):
    form = await request.form()
    verify_csrf(request, form)
    product = Product()
    try:
        await apply_product_form(db, product, form)
        db.add(product)
        db.flush()  # Flush to get product.id
        
        # Handle variants after product is saved
        handle_product_variants(db, product.id, form)
        
        db.commit()
    except ValueError as exc:
        db.rollback()
        categories = db.scalars(select(Category).where(Category.is_active.is_(True)).order_by(Category.name)).all()
        return templates.TemplateResponse(
            request=request,
            name="admin/products/form.html",
            context=template_context(
                request,
                admin=admin,
                product=None,
                categories=categories,
                form_values=product_form_values(form),
                error=str(exc),
            ),
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
        )
    set_flash(request, "Product created successfully.")
    return RedirectResponse("/admin/products", status_code=status.HTTP_303_SEE_OTHER)


@router.get("/products/{product_id}/edit")
def product_edit_page(
    request: Request,
    product_id: int,
    db: Session = Depends(get_db),
    admin: Admin = Depends(require_admin),
):
    product = db.get(Product, product_id)
    if not product:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Product not found")
    
    # Load variants for the product
    variants = db.scalars(
        select(ProductVariant).where(ProductVariant.product_id == product_id)
    ).all()
    
    categories = db.scalars(select(Category).where(Category.is_active.is_(True)).order_by(Category.name)).all()
    return templates.TemplateResponse(
        request=request,
        name="admin/products/form.html",
        context=template_context(
            request, 
            admin=admin, 
            product=product, 
            categories=categories, 
            variants=variants,  # Pass variants to template
            form_values={}, 
            error=None
        ),
    )


@router.post("/products/{product_id}/edit")
async def product_update(
    request: Request,
    product_id: int,
    db: Session = Depends(get_db),
    admin: Admin = Depends(require_admin),
):
    product = db.get(Product, product_id)
    if not product:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Product not found")
    form = await request.form()
    verify_csrf(request, form)
    try:
        await apply_product_form(db, product, form)
        
        # Handle variants update
        handle_product_variants(db, product.id, form)
        
        db.commit()
    except ValueError as exc:
        db.rollback()
        categories = db.scalars(select(Category).where(Category.is_active.is_(True)).order_by(Category.name)).all()
        variants = db.scalars(
            select(ProductVariant).where(ProductVariant.product_id == product_id)
        ).all()
        return templates.TemplateResponse(
            request=request,
            name="admin/products/form.html",
            context=template_context(
                request,
                admin=admin,
                product=product,
                categories=categories,
                variants=variants,
                form_values=product_form_values(form),
                error=str(exc),
            ),
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
        )
    set_flash(request, "Product updated successfully.")
    return RedirectResponse("/admin/products", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/products/{product_id}/archive")
async def product_archive(
    request: Request,
    product_id: int,
    db: Session = Depends(get_db),
    _: Admin = Depends(require_admin),
):
    form = await request.form()
    verify_csrf(request, form)
    product = db.get(Product, product_id)
    if product:
        product.is_active = False
        db.commit()
        set_flash(request, "Product archived. Existing order history remains safe.")
    return RedirectResponse("/admin/products", status_code=status.HTTP_303_SEE_OTHER)


@router.get("/categories")
def categories_page(
    request: Request,
    db: Session = Depends(get_db),
    admin: Admin = Depends(require_admin),
):
    categories = db.scalars(select(Category).order_by(Category.name)).all()
    return templates.TemplateResponse(
        request=request,
        name="admin/categories/list.html",
        context=template_context(request, admin=admin, categories=categories),
    )


@router.post("/categories")
async def category_create(
    request: Request,
    db: Session = Depends(get_db),
    _: Admin = Depends(require_admin),
):
    form = await request.form()
    verify_csrf(request, form)
    name = str(form.get("name", "")).strip()
    if len(name) < 2:
        set_flash(request, "Category name is too short.", "error")
    elif db.scalar(select(Category).where(func.lower(Category.name) == name.lower())):
        set_flash(request, "This category already exists.", "error")
    else:
        db.add(Category(name=name, slug=unique_category_slug(db, name)))
        db.commit()
        set_flash(request, "Category created successfully.")
    return RedirectResponse("/admin/categories", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/categories/{category_id}/toggle")
async def category_toggle(
    request: Request,
    category_id: int,
    db: Session = Depends(get_db),
    _: Admin = Depends(require_admin),
):
    form = await request.form()
    verify_csrf(request, form)
    category = db.get(Category, category_id)
    if category:
        category.is_active = not category.is_active
        db.commit()
    return RedirectResponse("/admin/categories", status_code=status.HTTP_303_SEE_OTHER)


@router.get("/orders")
def orders_list(
    request: Request,
    order_status: str = "",
    db: Session = Depends(get_db),
    admin: Admin = Depends(require_admin),
):
    query = select(Order).order_by(Order.created_at.desc())
    if order_status:
        try:
            query = query.where(Order.status == OrderStatus(order_status))
        except ValueError:
            order_status = ""
    orders = db.scalars(query).all()
    return templates.TemplateResponse(
        request=request,
        name="admin/orders/list.html",
        context=template_context(
            request,
            admin=admin,
            orders=orders,
            statuses=list(OrderStatus),
            selected_status=order_status,
        ),
    )


@router.get("/orders/{order_id}")
def order_detail(
    request: Request,
    order_id: int,
    db: Session = Depends(get_db),
    admin: Admin = Depends(require_admin),
):
    order = db.scalar(
        select(Order).options(joinedload(Order.items)).where(Order.id == order_id)
    )
    if not order:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Order not found")
    return templates.TemplateResponse(
        request=request,
        name="admin/orders/detail.html",
        context=template_context(
            request, admin=admin, order=order, statuses=list(OrderStatus)
        ),
    )


@router.post("/orders/{order_id}/status")
async def order_status_update(
    request: Request,
    order_id: int,
    db: Session = Depends(get_db),
    _: Admin = Depends(require_admin),
):
    form = await request.form()
    verify_csrf(request, form)
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Order not found")
    try:
        order.status = OrderStatus(str(form.get("status", "")))
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid status") from exc
    db.commit()
    set_flash(request, "Order status updated.")
    return RedirectResponse(
        f"/admin/orders/{order.id}", status_code=status.HTTP_303_SEE_OTHER
    )


# ============================================
# ✅ BULK IMPORT ROUTES
# ============================================

from openpyxl import load_workbook  # ✅ Import inside (after dependencies)


def _find_col(headers, possible_names):
    """Find column index by checking possible names."""
    for name in possible_names:
        normalized = name.lower().replace(" ", "").replace("_", "")
        if normalized in headers:
            return headers[normalized]
    return None


def _parse_price(value):
    """Parse price from various formats."""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _parse_int(value):
    """Parse integer from various formats."""
    if value is None:
        return 0
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return 0


def _generate_slug(name):
    """Generate URL slug from name."""
    slug = name.lower().strip()
    # Remove special characters
    import re
    slug = re.sub(r'[^a-z0-9]+', '-', slug)
    slug = slug.strip('-')
    return slug


def _get_or_create_category(db: Session, name: str) -> Category:
    """Get existing category or create new one."""
    if not name or name.strip() == "":
        name = "Uncategorized"
    
    cat = db.query(Category).filter(Category.name.ilike(name)).first()
    if cat:
        return cat
    
    slug = _generate_slug(name)
    existing = db.query(Category).filter(Category.slug == slug).first()
    if existing:
        slug = f"{slug}-{datetime.now().strftime('%H%M%S')}"
    
    cat = Category(name=name, slug=slug, is_active=True)
    db.add(cat)
    db.commit()
    db.refresh(cat)
    return cat


@router.get("/bulk-import")
def bulk_import_page(
    request: Request,
    admin: Admin = Depends(require_admin),
):
    """Show bulk import page."""
    # Get results from session if available
    last_results = request.session.pop("last_import_results", None)
    
    return templates.TemplateResponse(
        request=request,
        name="admin/bulk_import.html",
        context=template_context(
            request,
            admin=admin,
            last_results=last_results,
        ),
    )


@router.post("/bulk-import")
async def bulk_import(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin: Admin = Depends(require_admin),
):
    """Process bulk import Excel file."""
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        set_flash(request, "Please upload a valid Excel file (.xlsx or .xls)", "error")
        return RedirectResponse("/admin/bulk-import", status_code=status.HTTP_303_SEE_OTHER)
    
    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
        
        # Detect headers
        headers = {}
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            if row and any(cell is not None for cell in row):
                for idx, cell in enumerate(row):
                    if cell:
                        key = str(cell).strip().lower().replace(" ", "").replace("_", "")
                        headers[key] = idx
                break
        
        # Column mapping
        col_name = _find_col(headers, ["productname", "name", "title", "product"])
        col_price = _find_col(headers, ["price", "rate", "saleprice", "unitprice"])
        col_compare = _find_col(headers, ["compareatprice", "originalprice", "mrp", "oldprice", "discountprice"])
        col_stock = _find_col(headers, ["stock", "quantity", "qty", "inventory"])
        col_category = _find_col(headers, ["category", "department", "type", "group"])
        col_unit = _find_col(headers, ["unit", "size", "pack"])
        
        if col_name is None or col_price is None:
            set_flash(request, "Excel file must have 'Product Name' and 'Price' columns", "error")
            return RedirectResponse("/admin/bulk-import", status_code=status.HTTP_303_SEE_OTHER)
        
        results = {"updated": 0, "new": 0, "skipped": 0, "errors": []}
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or row[0] is None:
                    continue
                
                name = str(row[col_name]).strip() if col_name < len(row) and row[col_name] is not None else None
                if not name or name.lower() in ["productname", "name", ""]:
                    results["skipped"] += 1
                    continue
                
                # Parse values
                price = _parse_price(row[col_price]) if col_price is not None and col_price < len(row) else None
                compare_price = _parse_price(row[col_compare]) if col_compare is not None and col_compare < len(row) else None
                stock = _parse_int(row[col_stock]) if col_stock is not None and col_stock < len(row) else 0
                category_name = str(row[col_category]).strip() if col_category is not None and col_category < len(row) and row[col_category] is not None else "Uncategorized"
                unit = str(row[col_unit]).strip() if col_unit is not None and col_unit < len(row) and row[col_unit] is not None else "piece"
                
                # Find existing product by name
                product = db.query(Product).filter(Product.name.ilike(name)).first()
                
                if product:
                    # UPDATE existing
                    product.price = Decimal(str(price)) if price else product.price
                    product.compare_at_price = Decimal(str(compare_price)) if compare_price else product.compare_at_price
                    product.stock_quantity = max(0, stock)
                    product.stock_status = StockStatus.IN_STOCK if stock > 0 else StockStatus.OUT_OF_STOCK
                    product.unit = unit if unit else product.unit
                    results["updated"] += 1
                else:
                    # CREATE new
                    category = _get_or_create_category(db, category_name)
                    
                    slug = _generate_slug(name)
                    existing = db.query(Product).filter(Product.slug == slug).first()
                    if existing:
                        slug = f"{slug}-{datetime.now().strftime('%H%M%S')}"
                    
                    new_product = Product(
                        name=name,
                        slug=slug,
                        description=f"Bulk imported on {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                        price=Decimal(str(price)) if price else Decimal("0"),
                        compare_at_price=Decimal(str(compare_price)) if compare_price else None,
                        stock_quantity=max(0, stock),
                        stock_status=StockStatus.IN_STOCK if stock > 0 else StockStatus.OUT_OF_STOCK,
                        unit=unit if unit else "piece",
                        image_path=None,
                        is_active=True,
                        is_featured=False,
                        category_id=category.id,
                    )
                    db.add(new_product)
                    results["new"] += 1
            
            except Exception as e:
                results["errors"].append(f"Row {row_idx}: {str(e)}")
        
        db.commit()
        
        # Flash message with summary
        msg = f"✅ Import complete! New: {results['new']}, Updated: {results['updated']}, Skipped: {results['skipped']}"
        if results['errors']:
            msg += f" | ⚠️ Errors: {len(results['errors'])}"
        set_flash(request, msg, "success" if not results['errors'] else "warning")
        
        # Store detailed results in session for display
        request.session["last_import_results"] = results
        
    except Exception as e:
        db.rollback()
        set_flash(request, f"❌ Import failed: {str(e)}", "error")
    
    return RedirectResponse("/admin/bulk-import", status_code=status.HTTP_303_SEE_OTHER)