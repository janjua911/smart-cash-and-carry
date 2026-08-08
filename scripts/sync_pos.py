"""
Smart Cash & Carry — POS Excel to Website Sync Script

Yeh script karta kya hai:
1. stock.xlsx file read karta hai (POS se aayi hui)
2. Existing products ko update karta hai (price, stock, status)
3. Naye products ko automatically add kar deta hai
4. Categories ko auto-create karta hai agar nayi hon
5. Stock status auto-set hota hai (0 = out_of_stock, >0 = in_stock)
6. Images MANUALLY add karni hain baad mein admin se

Chalane ka tareeqa:
    venv\Scripts\activate
    python -m scripts.sync_pos

    Ya auto-watch ke liye:
    python -m scripts.sync_pos --watch
"""

import os
import sys
import time
import argparse
from datetime import datetime
from decimal import Decimal

from openpyxl import load_workbook
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

# Project root add karo taake imports kaam karein
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

from app.core.config import settings
from app.database import Base
from app.models.product import Product
from app.models.category import Category
from app.services.slug import generate_slug

# ============================================================
# CONFIGURATION — Yahan apna path set karo
# ============================================================

# Option A: Excel file project root mein rakho (easy)
EXCEL_FILE = r"D:\Hassan Afzal\CnC\centro\Smart-Cash-and-Carry-main\stock.xlsx"

# Option B: Agar POS kisi aur folder se export karti hai, yeh path change karo
# EXCEL_FILE = r"D:\POS\Exports\stock.xlsx"
# EXCEL_FILE = r"C:\Users\hassa\Desktop\stock.xlsx"

# Database connection
engine = create_engine(str(settings.DATABASE_URL))
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

# Excel column mapping (header names ko lowercase mein match karta hai)
COLUMN_MAP = {
    "id": ["productid", "id", "product_id", "item id", "itemid"],
    "name": ["productname", "name", "product name", "item name", "itemname", "title"],
    "category": ["category", "cat", "department", "type", "group"],
    "price": ["price", "rate", "cost", "sale price", "unit price"],
    "stock": ["stock", "quantity", "qty", "inventory", "units", "stock quantity"],
}


def normalize_header(value):
    """Header ko clean kar ke lowercase mein return karta hai"""
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "").replace("_", "")


def find_column_index(headers, possible_names):
    """Headers mein se best match dhundhta hai"""
    for idx, header in enumerate(headers):
        normalized = normalize_header(header)
        for name in possible_names:
            if normalize_header(name) == normalized:
                return idx
    return None


def get_or_create_category(db, category_name):
    """Category ko dhundo ya nayi banao"""
    if not category_name or str(category_name).strip() == "":
        category_name = "Uncategorized"
    
    category_name = str(category_name).strip()
    
    # Pehle exact match try karo
    cat = db.query(Category).filter(Category.name.ilike(category_name)).first()
    if cat:
        return cat
    
    # Nayi category banao
    slug = generate_slug(category_name)
    
    # Check ke slug unique hai
    existing = db.query(Category).filter(Category.slug == slug).first()
    if existing:
        slug = f"{slug}-{datetime.now().strftime('%H%M%S')}"
    
    cat = Category(
        name=category_name,
        slug=slug,
        is_active=True,
    )
    db.add(cat)
    db.commit()
    db.refresh(cat)
    print(f"    ➕ New category created: {category_name}")
    return cat


def sync_excel_to_db():
    """Main sync function"""
    
    if not os.path.exists(EXCEL_FILE):
        print(f"\n❌ ERROR: File not found: {EXCEL_FILE}")
        print("   stock.xlsx ko project root folder mein rakho ya EXCEL_FILE path change karo.")
        return False

    db = SessionLocal()
    try:
        print(f"\n{'='*60}")
        print(f"📊 SYNC STARTED: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"📁 File: {EXCEL_FILE}")
        print(f"{'='*60}")

        wb = load_workbook(EXCEL_FILE, data_only=True)
        ws = wb.active  # Pehli sheet

        # Headers read karo (pehli 5 rows check karte hain)
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            if row and any(cell is not None for cell in row):
                headers = [str(cell).strip() if cell is not None else "" for cell in row]
                break

        if not headers:
            print("❌ ERROR: Excel file mein headers nahi mile.")
            return False

        print(f"   Headers found: {headers}")

        # Column indexes dhundo
        col_id = find_column_index(headers, COLUMN_MAP["id"])
        col_name = find_column_index(headers, COLUMN_MAP["name"])
        col_category = find_column_index(headers, COLUMN_MAP["category"])
        col_price = find_column_index(headers, COLUMN_MAP["price"])
        col_stock = find_column_index(headers, COLUMN_MAP["stock"])

        print(f"   Columns mapped → ID:{col_id}, Name:{col_name}, Category:{col_category}, Price:{col_price}, Stock:{col_stock}")

        if col_name is None:
            print("❌ ERROR: 'Product Name' column nahi mili. Headers check karo.")
            return False

        stats = {"updated": 0, "new": 0, "skipped": 0, "errors": 0}

        # Data rows process karo (row 2 se start, assuming row 1 = header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or row[0] is None:
                    continue

                # Values nikalo
                raw_id = row[col_id] if col_id is not None and col_id < len(row) else None
                name = str(row[col_name]).strip() if col_name is not None and col_name < len(row) and row[col_name] is not None else None
                category_name = str(row[col_category]).strip() if col_category is not None and col_category < len(row) and row[col_category] is not None else "Uncategorized"
                price = row[col_price] if col_price is not None and col_price < len(row) else 0
                stock = row[col_stock] if col_stock is not None and col_stock < len(row) else 0

                if not name or name.lower() in ["productname", "name", ""]:
                    continue

                # Price aur stock ko clean karo
                try:
                    price = float(price) if price else 0.0
                except (ValueError, TypeError):
                    price = 0.0
                
                try:
                    stock = int(float(stock)) if stock is not None else 0
                except (ValueError, TypeError):
                    stock = 0

                # Product dhundo (pehle ID se, phir name se)
                product = None
                if raw_id:
                    try:
                        pid = int(float(raw_id))
                        product = db.query(Product).filter(Product.id == pid).first()
                    except (ValueError, TypeError):
                        pass
                
                if not product:
                    product = db.query(Product).filter(Product.name.ilike(name)).first()

                if product:
                    # ===== EXISTING PRODUCT — UPDATE =====
                    old_price = float(product.price)
                    old_stock = product.stock_quantity
                    old_status = product.stock_status.value if product.stock_status else "unknown"

                    product.price = Decimal(str(price)) if price else product.price
                    product.stock_quantity = max(0, stock)
                    
                    # Auto status
                    if product.stock_quantity > 0:
                        product.stock_status = "in_stock"
                    else:
                        product.stock_status = "out_of_stock"

                    changed = (old_price != float(product.price) or 
                               old_stock != product.stock_quantity or 
                               old_status != product.stock_status.value)

                    if changed:
                        stats["updated"] += 1
                        print(f"    🔄 UPDATED: {name}")
                        print(f"       Price: {old_price} → {product.price} | Stock: {old_stock} → {product.stock_quantity} | Status: {old_status} → {product.stock_status.value}")
                    else:
                        stats["skipped"] += 1

                else:
                    # ===== NEW PRODUCT — CREATE =====
                    category = get_or_create_category(db, category_name)
                    
                    slug = generate_slug(name)
                    # Slug unique check
                    existing_slug = db.query(Product).filter(Product.slug == slug).first()
                    if existing_slug:
                        slug = f"{slug}-{datetime.now().strftime('%H%M%S')}"

                    new_product = Product(
                        name=name,
                        slug=slug,
                        description=f"Auto-imported from POS on {datetime.now().strftime('%Y-%m-%d')}. Image to be added manually.",
                        price=Decimal(str(price)) if price else Decimal("0"),
                        compare_at_price=None,
                        stock_quantity=max(0, stock),
                        stock_status="in_stock" if stock > 0 else "out_of_stock",
                        unit="piece",
                        image_path=None,  # ⚠️ Manual add karni hai
                        is_active=True,
                        is_featured=False,
                        category_id=category.id,
                    )
                    db.add(new_product)
                    stats["new"] += 1
                    print(f"    ➕ NEW PRODUCT: {name}")
                    print(f"       Price: Rs. {price} | Stock: {stock} | Category: {category_name}")
                    print(f"       ⚠️  Image NOT set — add manually from admin panel")

            except Exception as e:
                stats["errors"] += 1
                print(f"    ❌ ERROR on row {row_idx}: {e}")

        db.commit()
        
        print(f"\n{'='*60}")
        print(f"✅ SYNC COMPLETE!")
        print(f"   🔄 Updated: {stats['updated']}")
        print(f"   ➕ New:     {stats['new']}")
        print(f"   ⏭️  Skipped: {stats['skipped']}")
        print(f"   ❌ Errors:  {stats['errors']}")
        print(f"{'='*60}\n")
        return True

    except Exception as e:
        db.rollback()
        print(f"\n❌ FATAL ERROR: {e}")
        return False
    finally:
        db.close()


# ==================== AUTO-WATCH MODE ====================

class ExcelSyncHandler(FileSystemEventHandler):
    def __init__(self):
        self.last_processed = 0

    def on_modified(self, event):
        if event.src_path.endswith((".xlsx", ".xls")):
            now = time.time()
            if now - self.last_processed > 3:  # 3 second debounce
                self.last_processed = now
                time.sleep(1)
                print(f"\n📁 File change detected: {event.src_path}")
                sync_excel_to_db()


def watch_mode():
    """Background mein file watch karta hai"""
    print(f"\n👁️  WATCH MODE STARTED")
    print(f"   Watching: {EXCEL_FILE}")
    print(f"   stock.xlsx save karo — auto-sync ho jayega!")
    print(f"   Rokne ke liye: Ctrl+C\n")

    # Pehli dafa sync
    sync_excel_to_db()

    handler = ExcelSyncHandler()
    observer = Observer()
    watch_path = os.path.dirname(os.path.abspath(EXCEL_FILE)) or "."
    observer.schedule(handler, path=watch_path, recursive=False)
    observer.start()

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("\n\n🛑 Stopping watcher...")
        observer.stop()
    observer.join()


# ==================== MAIN ====================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="POS to Website Sync")
    parser.add_argument("--watch", action="store_true", help="Auto-watch mode (background)")
    args = parser.parse_args()

    if args.watch:
        watch_mode()
    else:
        sync_excel_to_db()